"""
services/customer_import/parser.py
──────────────────────────────────
Pure-function CSV / XLSX parser used by the import wizard.

Returns a `ParsedFile` with:
    - headers:  ordered list of column names (deduped, stripped)
    - rows:     list of dicts {header: cell_value} — strings only
    - kind:     "csv" or "xlsx"
    - filename: original filename (for audit)

Hard limits keep the wizard responsive and prevent abuse:
    MAX_ROWS    = 10,000
    MAX_BYTES   = 5 MB

The parser never touches the DB and never normalizes data — that is
the normalizer's job. CSV detection accepts comma, semicolon, or
tab separators (sniffed automatically).
"""
from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from typing import List, Optional

logger = logging.getLogger("nahla.import.parser")

MAX_ROWS  = 10_000
MAX_BYTES = 5 * 1024 * 1024  # 5 MB

# Encodings tried in order for CSV — Excel-on-Windows users often
# export UTF-8 with BOM; older Arabic exports may be cp1256.
_CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1256", "windows-1252")


class ParseError(ValueError):
    """Raised when the uploaded file cannot be parsed at all (corrupt,
    empty, wrong format)."""


@dataclass
class ParsedFile:
    headers:  List[str]
    rows:     List[dict] = field(default_factory=list)
    kind:     str = "csv"
    filename: str = ""

    @property
    def total_rows(self) -> int:
        return len(self.rows)


# ── Public API ───────────────────────────────────────────────────────────────

def parse_upload(
    *,
    content: bytes,
    filename: str,
) -> ParsedFile:
    """Sniff the file type from the filename + bytes and dispatch to
    the right parser. Raises `ParseError` on anything unrecoverable."""
    if not content:
        raise ParseError("الملف فارغ.")
    if len(content) > MAX_BYTES:
        raise ParseError(
            f"حجم الملف يتجاوز الحد المسموح ({MAX_BYTES // (1024*1024)} ميجابايت)."
        )

    name = (filename or "").lower().strip()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content, filename)
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return _parse_csv(content, filename)

    # Fallback sniffing: openpyxl files start with PK (zip header).
    if content[:2] == b"PK":
        return _parse_xlsx(content, filename)
    return _parse_csv(content, filename)


# ── CSV ──────────────────────────────────────────────────────────────────────

def _decode_csv(content: bytes) -> str:
    last_err: Optional[Exception] = None
    for enc in _CSV_ENCODINGS:
        try:
            return content.decode(enc)
        except UnicodeDecodeError as exc:
            last_err = exc
            continue
    raise ParseError(f"تعذّر قراءة ترميز الملف ({last_err}).")


def _parse_csv(content: bytes, filename: str) -> ParsedFile:
    text = _decode_csv(content)
    if not text.strip():
        raise ParseError("الملف فارغ.")

    # Sniff the delimiter from a generous sample so we handle comma /
    # semicolon (very common in Arabic Excel exports) / tab files.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # safe default

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    try:
        header_row = next(reader)
    except StopIteration:
        raise ParseError("لا توجد رؤوس أعمدة في الملف.") from None

    headers = _dedupe_headers([_clean_cell(h) for h in header_row])
    rows: List[dict] = []
    for row in reader:
        if len(rows) >= MAX_ROWS:
            logger.warning("CSV truncated at %s rows", MAX_ROWS)
            break
        if not any(_clean_cell(c) for c in row):
            continue  # skip fully empty rows
        record = {}
        for idx, header in enumerate(headers):
            value = _clean_cell(row[idx]) if idx < len(row) else ""
            record[header] = value
        rows.append(record)

    if not rows:
        raise ParseError("لم يتم العثور على أي سطر بيانات بعد الرؤوس.")
    if not headers:
        raise ParseError("لم يتم العثور على رؤوس أعمدة صالحة.")

    return ParsedFile(
        headers=headers, rows=rows, kind="csv", filename=filename or "",
    )


# ── XLSX ─────────────────────────────────────────────────────────────────────

def _parse_xlsx(content: bytes, filename: str) -> ParsedFile:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - install-time guard
        raise ParseError(
            "حزمة openpyxl غير مثبّتة على الخادم — لا يمكن قراءة XLSX."
        ) from exc

    try:
        wb = load_workbook(
            filename=io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ParseError(f"تعذّر فتح ملف XLSX: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise ParseError("لم يتم العثور على ورقة عمل في الملف.")

    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        raise ParseError("لا توجد رؤوس أعمدة في الملف.") from None

    headers = _dedupe_headers([_clean_cell(h) for h in (header_row or ())])
    rows: List[dict] = []
    for row in iterator:
        if len(rows) >= MAX_ROWS:
            logger.warning("XLSX truncated at %s rows", MAX_ROWS)
            break
        cells = [_clean_cell(c) for c in (row or ())]
        if not any(cells):
            continue
        record = {}
        for idx, header in enumerate(headers):
            value = cells[idx] if idx < len(cells) else ""
            record[header] = value
        rows.append(record)

    try:
        wb.close()
    except Exception:  # noqa: silent-ok — best-effort workbook cleanup; leak doesn't affect import correctness
        pass

    if not headers:
        raise ParseError("لم يتم العثور على رؤوس أعمدة صالحة.")
    if not rows:
        raise ParseError("لم يتم العثور على أي سطر بيانات بعد الرؤوس.")

    return ParsedFile(
        headers=headers, rows=rows, kind="xlsx", filename=filename or "",
    )


# ── Helpers ──────────────────────────────────────────────────────────────────

def _clean_cell(value) -> str:
    if value is None:
        return ""
    # openpyxl returns native types — coerce to str for the import stage.
    if isinstance(value, float) and value.is_integer():
        # Avoid "5421234567.0" from numeric phone columns.
        return str(int(value))
    return str(value).strip()


def _dedupe_headers(headers: List[str]) -> List[str]:
    """Excel files sometimes contain blank or duplicated header cells.
    Replace blanks with `column_<n>` and disambiguate duplicates."""
    seen = {}
    out: List[str] = []
    for idx, raw in enumerate(headers):
        name = raw or f"column_{idx + 1}"
        base = name
        n = seen.get(base, 0)
        if n:
            name = f"{base} ({n + 1})"
        seen[base] = n + 1
        out.append(name)
    return out
